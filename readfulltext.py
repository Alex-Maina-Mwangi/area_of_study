#This script scrapes through europepmc and reads json data from open access scientific articles and extracts the country where the research was done.
#Written by Alex Maina Mwangi - Librarian Kenya Medical Research Institute/Wellcome Trust Research Programme - Centre for Geography Medicine Research - Coast
#Kilifi Kenya

import requests
from bs4 import BeautifulSoup
import re
import geonamescache
import spacy
from spacy.matcher import PhraseMatcher
import pycountry
from openpyxl import Workbook
'''
pmcids = ['PMC10952650','PMC7616763','PMC10808817','PMC7616760','PMC10808817','PMC10842664',
'PMC10716622','PMC10716626','PMC10716625','PMC10716621','PMC10765698','PMC10773318',
'PMC10806884','PMC10900964','PMC10900966','PMC10800037','PMC10849973','PMC10919173',
'PMC10823747','PMC10864189','PMC10840034','PMC10846728','PMC10864294','PMC7616643',
'PMC11024825','PMC10927117','PMC10852234','PMC10923414','PMC11031140','PMC7616644',
'PMC10867106','PMC10854052','PMC10877892','PMC11005831','PMC10906885','PMC10905788',
'PMC10901986','PMC10871522','PMC10921557','PMC11289910','PMC10928745','PMC10808817',
'PMC10716622','PMC10852234','PMC10906885','PMC10892224','PMC10923414','PMC10935770',
'PMC11126395','PMC10937349','PMC10949203','PMC10945606','PMC10946148','PMC10949706',
'PMC10950691','PMC10977730','PMC10977907','PMC10996164','PMC11004247','PMC11017571',
'PMC11059186','PMC11064105','PMC11069174','PMC11289976','PMC11046809','PMC11027528',
'PMC11033408','PMC11021640','PMC11043685','PMC10832137','PMC11039628','PMC11076563',
'PMC10902387','PMC11289976','PMC11094066','PMC11287631','PMC11149845','PMC11087563',
'PMC11103839','PMC11106534','PMC7616653','PMC11088304','PMC11097591','PMC7616646',
'PMC11139289','PMC11110293','PMC11106525','PMC11097862','PMC11167876','PMC11189180',
'PMC11192416','PMC11216587','PMC11184184','PMC7616761','PMC11253580','PMC11163641',
'PMC7616506','PMC11229757','PMC11197641','PMC11232189','PMC11253363','PMC11253331',
'PMC7616119','PMC11254782','PMC11216554','PMC11249019','PMC11214617','PMC11216563',
'PMC11232288','PMC11257704','PMC11549274','PMC11293692','PMC11365168','PMC11293714',
'PMC11293042','PMC11342035','PMC11404254','PMC11409615','PMC11464403','PMC11349500',
'PMC11338195','PMC11337709','PMC11329769','PMC11328904','PMC11345449','PMC11382623',
'PMC11331881','PMC11306574','PMC11300237','PMC11543637','PMC11349107','PMC11542533',
'PMC11375402','PMC11357821','PMC11367973','PMC11387285','PMC11370027','PMC11410205',
'PMC11375914','PMC11392261','PMC11450608','PMC11403289','PMC11399766','PMC11525066',
'PMC11438083','PMC11462847','PMC11542398','PMC7616702','PMC11403289','PMC11536953',
'PMC11436855','PMC11410205','PMC11438207','PMC11502999','PMC11530017','PMC11441138',
'PMC11499763','PMC11519284','PMC11461663','PMC11464310','PMC11497170','PMC11515546',
'PMC11590815','PMC11603168','PMC11530017','PMC11552008','PMC11581349','PMC11605845',
'PMC11606115','PMC11622353','PMC11589854','PMC11613948','PMC11624724','PMC11618498',
'PMC11669216','PMC11654438','PMC11131180','PMC11682625','PMC7617250','PMC11675204',
'PMC11638174','PMC11285936']
'''
pmcids = ['PMC11525066','PMC10864294','PMC11064105','PMC11438083','PMC11069174','PMC10716622','PMC11349107',
'PMC11338195','PMC11718159','PMC11285936','PMC11289976','PMC11723863','PMC11462847','PMC10950691',
'PMC10996164','PMC10919173','PMC11345449','PMC11542398','PMC11562111','PMC11094066','PMC11375402',
'PMC11549274','PMC11287631','PMC11149845','PMC10927117','PMC7616702','PMC10852234','PMC10952650',
'PMC11357821','PMC11404254','PMC10923414','PMC11637338','PMC11367973','PMC11087563','PMC11622353',
'PMC11046809','PMC11293692','PMC10867106','PMC11103839','PMC11654438','PMC10854052','PMC11192416',
'PMC11106534','PMC10877892','PMC10823747','PMC11229757','PMC11409615','PMC10937349','PMC7616653',
'PMC11197641','PMC11464403','PMC10765698','PMC11420681','PMC11088304','PMC11131180','PMC10977907',
'PMC11387285','PMC10832137','PMC11536953','PMC11613948','PMC11624724','PMC10864189','PMC11342035',
'PMC11232189','PMC11097591','PMC11005831','PMC11590815','PMC10906885','PMC11365168','PMC11370027',
'PMC10905788','PMC10840034','PMC11253331','PMC11436855','PMC11603168','PMC7616119','PMC11216587',
'PMC11410205','PMC11438207','PMC11737602','PMC10892224','PMC11399766','PMC11254782','PMC11530017',
'PMC10901986','PMC10871522','PMC11441138','PMC11618498','PMC10832587','PMC11499763','PMC10945606',
'PMC11375914','PMC11293042','PMC10846728','PMC11139289','PMC7617250','PMC11382623','PMC11675204',
'PMC10949706','PMC10773318','PMC10977730','PMC11552008','PMC11329769','PMC11461663','PMC11581349',
'PMC10935770','PMC10921557','PMC11289910','PMC11337709','PMC11106525','PMC11392261','PMC11605845',
'PMC11097862','PMC10928745','PMC11253580','PMC10800037','PMC11581619','PMC11331881','PMC11163641',
'PMC11450608','PMC10849973','PMC10842664','PMC11464310','PMC11004247','PMC7616506','PMC11300237',
'PMC11515546','PMC10902387','PMC11349500','PMC11216563','PMC10900964']

#pmcids = ['PMC10952650','PMC7616763']

data = []
for i in pmcids:
    url = 'https://www.ebi.ac.uk/europepmc/webservices/rest/'+i+'/fullTextXML'
    #print(url)
    #url = 'https://www.ebi.ac.uk/europepmc/webservices/rest/PMC11285936/fullTextXML'
    #url = 'https://www.ebi.ac.uk/europepmc/webservices/rest/PMC10716622/fullTextXML'
    #url = 'https://www.ebi.ac.uk/europepmc/webservices/rest/PMC10952650/fullTextXML'
    #url = 'https://www.ebi.ac.uk/europepmc/webservices/rest/PMC10950691/fullTextXML'
    #url = 'https://www.ebi.ac.uk/europepmc/webservices/rest/PMC11525066/fullTextXML'
    xml_data = requests.get(url).content
    soup = BeautifulSoup(xml_data, 'xml')

    #get PMID
    pmid_tag = soup.find("article-id", {"pub-id-type": "pmid"})

    # Extract the PMID if not exist return None
    pmid = pmid_tag.text if pmid_tag else None

    # Print the PMID
    print("PMID:", pmid)

    # Extract the text content from the XML
    #text_content = soup.get_text()
    # Extract all <p> tags (text content)
    paragraphs = soup.find_all("p")

    # Combine text from all <p> tags
    text_content = " ".join(p.get_text() for p in paragraphs)
    gc = geonamescache.GeonamesCache()
    countries = gc.get_countries()
    country_names = [country['name'] for country in countries.values()]

    # Extract mentioned country names
    mentioned_countries = [country for country in country_names if country in text_content]


    '''
    # Get a list of all country names from pycountry
    country_names = {country.name for country in pycountry.countries}

    # Initialize a set to store detected country names
    detected_countries = set()

    # Check for country names in the text content
    for word in text_content.split():
        # Remove punctuation and check if it's a country
        clean_word = word.strip(",.()[]").capitalize()
        if clean_word in country_names:
            detected_countries.add(clean_word)
    '''
    # Print the results
    ##print("Countries found:", detected_countries)
    kenyan_counties = [ "Baringo", "Bomet", "Bungoma", "Busia", "Elgeyo-Marakwet", "Embu", 
                        "Garissa", "Homa Bay", "Isiolo", "Kajiado", "Kakamega", "Kericho", 
            "Kiambu", "Kilifi", "Kirinyaga", "Kisii", "Kisumu", "Kitui", "Kwale", 
            "Laikipia", "Lamu", "Machakos", "Makueni", "Mandera", "Marsabit", 
            "Meru", "Migori", "Mombasa", "Murang'a", "Nairobi", "Nakuru", 
            "Nandi", "Narok", "Nyamira", "Nyandarua", "Nyeri", "Samburu", 
            "Siaya", "Taita-Taveta", "Tana River", "Tharaka-Nithi", "Trans Nzoia", 
            "Turkana", "Uasin Gishu", "Vihiga", "Wajir", "West Pokot"
                ]

    # Load spaCy's English language model
    nlp = spacy.load("en_core_web_sm")

    # Create a PhraseMatcher and add counties
    matcher = PhraseMatcher(nlp.vocab)
    patterns = [nlp.make_doc(county) for county in kenyan_counties]
    matcher.add("KenyanCounties", patterns)

    def find_counties_spacy(text, matcher):
        doc = nlp(text)
        matches = matcher(doc)
        found_counties = set([doc[start:end].text for match_id, start, end in matches])
        return found_counties

    # Example Usage
    #text = "Garissa and Nairobi are important economic hubs."
    found_counties = find_counties_spacy(text_content, matcher)
    ##print("Found Counties:", found_counties)
    output = [[pmid],mentioned_countries,found_counties,[url]]
    #print(output)
    #country_names = output[0]
    #county_names = output[1]
    #print(county_names)
    data.append(output)

print(data)
# Create a new Excel workbook
workbook = Workbook()
sheet = workbook.active
sheet.title = "List Data"

# Add column headers
sheet.cell(row=1, column=1, value="PMID")
sheet.cell(row=1, column=2, value="Countries")
sheet.cell(row=1, column=3, value="Counties")
sheet.cell(row=1, column=4, value="url")

# Write the data to Excel starting from the second row
row = 2
for record in data:
    col = 1
    for sublist in record:
        cleaned_sublist = [str(item) if item is not None else "" for item in sublist]
 
        sheet.cell(row=row, column=col, value=", ".join(cleaned_sublist))
        col += 1
    row += 1

# Save the workbook
output_file = "newest_list_data_with_headers.xlsx"
workbook.save(output_file)

print(f"Data with headers saved to {output_file}!")
